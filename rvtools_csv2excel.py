#!/usr/bin/env python3
"""
RVTools CSV to Excel Converter

This tool converts RVTools CSV exports to Excel format, preserving data and applying
formatting to match the official RVTools Excel export format.

Features:
- Convert single or multiple CSV files to a single Excel workbook
- Scan directories for CSV files
- Apply RVTools-style formatting
- Preserve data types
- Reorder sheets to match standard RVTools order
- Add metadata sheet

Usage:
    python rvtools_csv2excel.py [options]

Options:
    -h, --help              Show this help message and exit
    -i, --input DIR         Input directory containing CSV files (default: current directory)
    -o, --output FILE       Output Excel file (default: rvtools_export.xlsx)
    -r, --recursive         Scan subdirectories for CSV files
    -p, --prefix PREFIX     Only process files with this prefix (default: RVTools_tab)
    -v, --verbose           Show detailed processing information
"""

import os
import sys
import argparse
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime
import re
import traceback

# Define standard RVTools sheet order
STANDARD_SHEET_ORDER = [
    'vInfo', 'vCPU', 'vMemory', 'vDisk', 'vPartition', 'vNetwork', 'vCD', 'vUSB', 
    'vSnapshot', 'vTools', 'vSource', 'vRP', 'vCluster', 'vHost', 'vHBA', 'vNIC', 
    'vSwitch', 'vPort', 'dvSwitch', 'dvPort', 'vSC_VMK', 'vDatastore', 'vMultiPath', 
    'vLicense', 'vFileInfo', 'vHealth', 'vMetaData'
]

def parse_arguments():
    """Parse command line arguments."""
    parser = argparse.ArgumentParser(description='Convert RVTools CSV exports to Excel format')
    parser.add_argument('-i', '--input', default='.', help='Input directory containing CSV files')
    parser.add_argument('-o', '--output', default='rvtools_export.xlsx', help='Output Excel file')
    parser.add_argument('-r', '--recursive', action='store_true', help='Scan subdirectories for CSV files')
    parser.add_argument('-p', '--prefix', default='RVTools_tab', help='Only process files with this prefix')
    parser.add_argument('-v', '--verbose', action='store_true', help='Show detailed processing information')
    
    return parser.parse_args()

def find_csv_files(directory, recursive=False, prefix=None, verbose=False):
    """Find CSV files in the specified directory."""
    csv_files = []
    
    if verbose:
        print(f"Scanning for CSV files in {directory}" + (" and subdirectories" if recursive else ""))
    
    if recursive:
        for root, _, files in os.walk(directory):
            for file in files:
                if file.lower().endswith('.csv'):
                    if prefix is None or file.startswith(prefix):
                        csv_files.append(os.path.join(root, file))
    else:
        for file in os.listdir(directory):
            if file.lower().endswith('.csv'):
                if prefix is None or file.startswith(prefix):
                    csv_files.append(os.path.join(directory, file))
    
    if verbose:
        print(f"Found {len(csv_files)} CSV files")
    
    return csv_files

def get_sheet_name_from_filename(filename, prefix='RVTools_tab'):
    """Extract a suitable sheet name from the CSV filename."""
    base_name = os.path.basename(filename)
    
    # If the file follows the RVTools naming convention (RVTools_tab{SheetName}.csv)
    if base_name.startswith(prefix):
        sheet_name = base_name[len(prefix):].replace('.csv', '')
    else:
        # Otherwise just use the filename without extension
        sheet_name = os.path.splitext(base_name)[0]
    
    # Ensure sheet name is valid for Excel (max 31 chars, no special chars)
    sheet_name = re.sub(r'[\[\]:*?/\\]', '', sheet_name)  # Remove invalid chars
    sheet_name = sheet_name[:31]  # Truncate if too long
    
    return sheet_name

def apply_header_formatting(worksheet):
    """Apply formatting to the header row of the worksheet."""
    # Define header style (black background, white text, bold)
    header_font = Font(name='Verdana', size=9, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')
    header_alignment = Alignment(horizontal='left')
    
    # Apply formatting to header row
    for cell in worksheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

def apply_data_formatting(worksheet):
    """Apply formatting to the data rows of the worksheet."""
    # Define data row style
    data_font = Font(name='Verdana', size=9)
    data_alignment = Alignment(horizontal='left')
    
    # Apply formatting to data rows
    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.font = data_font
            cell.alignment = data_alignment
            
            # Apply conditional formatting for specific columns
            column_name = worksheet.cell(row=1, column=cell.column).value
            
            # Format Powerstate column
            if column_name == 'Powerstate' and cell.value == 'poweredOff':
                cell.font = Font(name='Verdana', size=9, color='FF0000')  # Red text for powered off VMs
            
            # Format Config status column
            if column_name == 'Config status':
                if cell.value == 'green':
                    cell.font = Font(name='Verdana', size=9, color='008000')  # Green text
                elif cell.value == 'red':
                    cell.font = Font(name='Verdana', size=9, color='FF0000')  # Red text
                elif cell.value == 'yellow':
                    cell.font = Font(name='Verdana', size=9, color='FFA500')  # Orange text
            
            # Format boolean values consistently
            if isinstance(cell.value, bool):
                cell.value = str(cell.value)

def auto_adjust_column_width(worksheet):
    """Adjust column widths based on content."""
    for column in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        
        # Find the maximum content length in the column
        for cell in column:
            if cell.value:
                try:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
                except:
                    pass
        
        # Set the column width (with some padding)
        adjusted_width = max_length + 2
        if adjusted_width > 100:  # Cap the width for very long content
            adjusted_width = 100
        elif adjusted_width < 8:  # Minimum width
            adjusted_width = 8
        worksheet.column_dimensions[column_letter].width = adjusted_width

def clean_csv_data(csv_file, verbose=False):
    """Clean and prepare CSV data for Excel conversion."""
    try:
        # Try to detect the encoding and delimiter
        with open(csv_file, 'r', encoding='utf-8') as f:
            sample = f.read(4096)
        
        # Check if we need to handle quoted fields with commas
        has_quotes = '"' in sample
        
        # Read the CSV file with pandas
        if has_quotes:
            df = pd.read_csv(csv_file, encoding='utf-8', quotechar='"', escapechar='\\')
        else:
            df = pd.read_csv(csv_file, encoding='utf-8')
        
        # Clean column names (remove extra whitespace)
        df.columns = [col.strip() for col in df.columns]
        
        if verbose:
            print(f"  - Read {len(df)} rows and {len(df.columns)} columns")
        
        return df
    except Exception as e:
        if verbose:
            print(f"  - Error reading with UTF-8: {str(e)}")
        
        # Try with different encoding if UTF-8 fails
        try:
            if has_quotes:
                df = pd.read_csv(csv_file, encoding='latin1', quotechar='"', escapechar='\\')
            else:
                df = pd.read_csv(csv_file, encoding='latin1')
                
            df.columns = [col.strip() for col in df.columns]
            
            if verbose:
                print(f"  - Successfully read with latin1 encoding: {len(df)} rows")
            
            return df
        except Exception as e2:
            if verbose:
                print(f"  - Failed with alternative encoding: {str(e2)}")
            
            # Last resort: try with different quoting options
            try:
                df = pd.read_csv(csv_file, quoting=3)  # QUOTE_NONE
                df.columns = [col.strip() for col in df.columns]
                
                if verbose:
                    print(f"  - Successfully read with QUOTE_NONE: {len(df)} rows")
                
                return df
            except Exception as e3:
                if verbose:
                    print(f"  - All reading attempts failed: {str(e3)}")
                return None

def convert_csv_to_excel(csv_files, output_file, prefix='RVTools_tab', verbose=False):
    """Convert CSV files to Excel format with proper formatting."""
    # Create a new Excel workbook
    workbook = openpyxl.Workbook()
    
    # Remove the default sheet
    default_sheet = workbook.active
    workbook.remove(default_sheet)
    
    # Track processed sheets for ordering
    processed_sheets = []
    
    # Process each CSV file
    for csv_file in csv_files:
        try:
            if verbose:
                print(f"Processing {csv_file}")
            
            # Get sheet name from filename
            sheet_name = get_sheet_name_from_filename(csv_file, prefix)
            
            if verbose:
                print(f"  - Sheet name: {sheet_name}")
            
            # Clean and read CSV data
            df = clean_csv_data(csv_file, verbose)
            if df is None:
                if verbose:
                    print(f"  - Skipping file due to read errors")
                continue
            
            # Create a sheet
            worksheet = workbook.create_sheet(title=sheet_name)
            processed_sheets.append(sheet_name)
            
            # Write header
            for col_idx, column_name in enumerate(df.columns, start=1):
                cell = worksheet.cell(row=1, column=col_idx)
                cell.value = column_name
            
            # Write data
            for row_idx, row_data in df.iterrows():
                for col_idx, value in enumerate(row_data, start=1):
                    cell = worksheet.cell(row=row_idx + 2, column=col_idx)
                    cell.value = value
            
            # Apply formatting
            if verbose:
                print(f"  - Applying formatting")
            
            apply_header_formatting(worksheet)
            apply_data_formatting(worksheet)
            auto_adjust_column_width(worksheet)
            
            # Freeze the header row
            worksheet.freeze_panes = 'A2'
            
            if verbose:
                print(f"  - Completed processing")
        
        except Exception as e:
            if verbose:
                print(f"  - Error processing {csv_file}: {str(e)}")
                print(traceback.format_exc())
            else:
                print(f"Error processing {csv_file}: {str(e)}")
    
    # Add metadata sheet
    if verbose:
        print("Adding metadata sheet")
    
    metadata_sheet = workbook.create_sheet(title='vMetaData')
    metadata_sheet['A1'] = 'RVTools major version'
    metadata_sheet['B1'] = 'RVTools version'
    metadata_sheet['C1'] = 'xlsx creation datetime'
    metadata_sheet['D1'] = 'Server'
    
    # Add metadata values
    metadata_sheet['A2'] = 4.4
    metadata_sheet['B2'] = '4.4.5.0'
    metadata_sheet['C2'] = datetime.now()
    metadata_sheet['D2'] = 'Converted by RVTools CSV2Excel Tool'
    
    # Reorder sheets to match RVTools standard order if possible
    if verbose:
        print("Reordering sheets")
    
    # Create a new order based on standard order and what we actually have
    new_order = []
    for sheet in STANDARD_SHEET_ORDER:
        if sheet in processed_sheets or sheet == 'vMetaData':
            new_order.append(sheet)
    
    # Add any sheets we have that aren't in the standard order
    for sheet in processed_sheets:
        if sheet not in new_order:
            new_order.append(sheet)
    
    # Add metadata at the end if not already included
    if 'vMetaData' not in new_order:
        new_order.append('vMetaData')
    
    # Reorder the sheets
    for i, sheet_name in enumerate(new_order):
        if sheet_name in workbook.sheetnames:
            workbook.move_sheet(sheet_name, i)
    
    # Save the workbook
    if verbose:
        print(f"Saving workbook to {output_file}")
    
    workbook.save(output_file)
    
    if verbose:
        print(f"Excel file created with {len(workbook.sheetnames)} sheets")
    
    return output_file

def main():
    """Main function to handle command line arguments and execute conversion."""
    # Parse command line arguments
    args = parse_arguments()
    
    # Find CSV files
    csv_files = find_csv_files(
        args.input, 
        recursive=args.recursive, 
        prefix=args.prefix if args.prefix else None,
        verbose=args.verbose
    )
    
    if not csv_files:
        print(f"No CSV files found in {args.input}")
        return
    
    print(f"Found {len(csv_files)} CSV files to convert")
    
    # Convert CSV files to Excel
    output_path = convert_csv_to_excel(
        csv_files, 
        args.output, 
        prefix=args.prefix,
        verbose=args.verbose
    )
    
    print(f"Conversion complete. Output file: {output_path}")

if __name__ == "__main__":
    main()